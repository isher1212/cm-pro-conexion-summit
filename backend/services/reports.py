import logging
import smtplib
import sqlite3
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from io import BytesIO
from typing import Any

import httpx

logger = logging.getLogger(__name__)


# ── Report log ─────────────────────────────────────────────────────────────────

def log_report(
    conn: sqlite3.Connection,
    report_type: str,
    channel: str,
    status: str,
    error_message: str = "",
) -> None:
    try:
        conn.execute(
            """INSERT INTO report_log (report_type, channel, status, sent_at, error_message)
               VALUES (?, ?, ?, ?, ?)""",
            (report_type, channel, status, datetime.now().isoformat(), error_message),
        )
        conn.commit()
    except Exception as e:
        logger.warning(f"Failed to log report: {e}")


def get_report_log(conn: sqlite3.Connection, limit: int = 50) -> list[dict]:
    cursor = conn.execute(
        "SELECT * FROM report_log ORDER BY sent_at DESC LIMIT ?", (limit,)
    )
    return [dict(row) for row in cursor.fetchall()]


# ── Email HTML builders ────────────────────────────────────────────────────────

def build_daily_email(
    articles: list[dict],
    trends: list[dict],
    anomalies: list[dict],
    tip: str = "",
) -> str:
    articles_html = ""
    for a in articles[:5]:
        articles_html += f"""
        <div style="border-left:3px solid #6366f1;padding:10px 14px;margin-bottom:12px;background:#fafafa;border-radius:0 8px 8px 0;">
          <a href="{a.get('url','#')}" style="font-weight:600;color:#1e293b;text-decoration:none;font-size:14px;">{a.get('title','')}</a>
          <div style="font-size:11px;color:#94a3b8;margin:3px 0;">{a.get('source','')} · {datetime.now().strftime('%d %b %Y')}</div>
          <p style="font-size:12px;color:#475569;margin:6px 0 4px;">{a.get('summary','')}</p>
          {f'<p style="font-size:11px;color:#10b981;font-style:italic;">✅ {a["relevance"]}</p>' if a.get("relevance") else ''}
        </div>"""

    trends_html = ""
    for t in trends[:1]:
        trends_html = f"""
        <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:14px;margin-bottom:12px;">
          <div style="font-weight:600;color:#92400e;font-size:13px;">🔥 {t.get('keyword','')}</div>
          <p style="font-size:12px;color:#78350f;margin:6px 0;">{t.get('how_to_apply','')}</p>
          {f'<p style="font-size:12px;color:#b45309;"><strong>💡 Idea:</strong> {t["post_idea"]}</p>' if t.get("post_idea") else ''}
        </div>"""

    anomaly_html = ""
    if anomalies:
        items = "".join(
            f"<li style='font-size:12px;color:#b91c1c;'>{a['platform']}: engagement {a.get('direction','bajó')} {abs(a.get('change_pct',0))}%</li>"
            for a in anomalies
        )
        anomaly_html = f"""
        <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px;margin-bottom:12px;">
          <strong style="color:#991b1b;">⚠️ Alerta de métricas</strong>
          <ul style="margin:6px 0 0 16px;padding:0;">{items}</ul>
        </div>"""

    tip_html = f"""
        <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:12px;">
          <strong style="color:#166534;">💡 Tip del día</strong>
          <p style="font-size:12px;color:#15803d;margin:4px 0 0;">{tip}</p>
        </div>""" if tip else ""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;margin:0;padding:20px;">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.1);">
    <div style="background:#6366f1;padding:24px;">
      <h1 style="color:#fff;margin:0;font-size:20px;">CM Pro · Resumen diario</h1>
      <p style="color:#c7d2fe;margin:4px 0 0;font-size:13px;">{datetime.now().strftime('%A, %d de %B de %Y')}</p>
    </div>
    <div style="padding:24px;">
      <h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">🔍 Noticias del día</h2>
      {articles_html or '<p style="color:#94a3b8;font-size:12px;">Sin noticias nuevas hoy.</p>'}
      <h2 style="color:#1e293b;font-size:15px;margin:20px 0 12px;">🔥 Tendencia del día</h2>
      {trends_html or '<p style="color:#94a3b8;font-size:12px;">Sin tendencias nuevas hoy.</p>'}
      {anomaly_html}
      {tip_html}
    </div>
    <div style="padding:16px 24px;background:#f8fafc;border-top:1px solid #e2e8f0;">
      <p style="font-size:11px;color:#94a3b8;margin:0;">Generado por CM Pro · Conexión Summit</p>
    </div>
  </div>
</body>
</html>"""


def build_weekly_email(
    articles: list[dict],
    trends: list[dict],
    metrics_summary: list[dict],
    proposals: list[dict],
    anomalies: list[dict],
    recommendations: str = "",
) -> str:
    articles_html = "".join(
        f'<li style="font-size:12px;color:#475569;margin-bottom:6px;"><a href="{a.get("url","#")}" style="color:#6366f1;font-weight:600;">{a.get("title","")}</a> — {a.get("summary","")[:100]}</li>'
        for a in articles[:5]
    )

    trends_html = "".join(
        f'<li style="font-size:12px;color:#475569;margin-bottom:6px;"><strong>{t.get("keyword","")}</strong> ({t.get("platform","")}): {t.get("how_to_apply","")}</li>'
        for t in trends[:5]
    )

    metrics_html = "".join(
        f'<tr><td style="padding:6px 8px;font-size:12px;color:#1e293b;">{m.get("platform","")}</td>'
        f'<td style="padding:6px 8px;text-align:right;font-size:12px;">{m.get("followers",0):,}</td>'
        f'<td style="padding:6px 8px;text-align:right;font-size:12px;">{m.get("engagement_rate",0)}%</td>'
        f'<td style="padding:6px 8px;text-align:right;font-size:12px;color:#94a3b8;">{m.get("week_label","")}</td></tr>'
        for m in metrics_summary
    )

    proposals_html = "".join(
        f'<li style="font-size:12px;color:#475569;margin-bottom:6px;"><strong>{p.get("topic","")}</strong> · {p.get("platform","")} · {p.get("suggested_date","")}</li>'
        for p in proposals[:5]
    )

    rec_html = f'<p style="font-size:12px;color:#475569;">{recommendations}</p>' if recommendations else ""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;margin:0;padding:20px;">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.1);">
    <div style="background:#6366f1;padding:24px;">
      <h1 style="color:#fff;margin:0;font-size:20px;">CM Pro · Informe semanal</h1>
      <p style="color:#c7d2fe;margin:4px 0 0;font-size:13px;">{datetime.now().strftime('Semana del %d de %B de %Y')}</p>
    </div>
    <div style="padding:24px;">
      <h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">📈 Métricas de la semana</h2>
      <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
        <thead><tr style="border-bottom:2px solid #e2e8f0;">
          <th style="text-align:left;font-size:11px;color:#94a3b8;padding:6px 8px;">Red</th>
          <th style="text-align:right;font-size:11px;color:#94a3b8;padding:6px 8px;">Seguidores</th>
          <th style="text-align:right;font-size:11px;color:#94a3b8;padding:6px 8px;">Engagement</th>
          <th style="text-align:right;font-size:11px;color:#94a3b8;padding:6px 8px;">Semana</th>
        </tr></thead>
        <tbody>{metrics_html or '<tr><td colspan="4" style="font-size:12px;color:#94a3b8;padding:8px;">Sin métricas esta semana.</td></tr>'}</tbody>
      </table>
      <h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">🔍 Noticias más relevantes</h2>
      <ul style="padding-left:16px;margin:0 0 20px;">{articles_html or '<li style="font-size:12px;color:#94a3b8;">Sin noticias.</li>'}</ul>
      <h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">🔥 Tendencias de la semana</h2>
      <ul style="padding-left:16px;margin:0 0 20px;">{trends_html or '<li style="font-size:12px;color:#94a3b8;">Sin tendencias.</li>'}</ul>
      <h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">📅 Propuesta de parrilla próxima semana</h2>
      <ul style="padding-left:16px;margin:0 0 20px;">{proposals_html or '<li style="font-size:12px;color:#94a3b8;">Sin propuestas aprobadas.</li>'}</ul>
      {'<h2 style="color:#1e293b;font-size:15px;margin:0 0 12px;">💬 Recomendaciones estratégicas</h2>' + rec_html if rec_html else ''}
    </div>
    <div style="padding:16px 24px;background:#f8fafc;border-top:1px solid #e2e8f0;">
      <p style="font-size:11px;color:#94a3b8;margin:0;">Generado por CM Pro · Conexión Summit</p>
    </div>
  </div>
</body>
</html>"""


# ── Telegram message builders ──────────────────────────────────────────────────

def build_telegram_intelligence_message(articles: list[dict]) -> str:
    lines = ["🔍 *Noticias del día — Conexión Summit*\n"]
    for a in articles[:5]:
        title = a.get("title", "Sin título")
        source = a.get("source", "")
        summary = a.get("summary", "")[:120]
        url = a.get("url", "")
        lines.append(f"• *{title}*\n  _{source}_ — {summary}\n  {url}\n")
    msg = "\n".join(lines)
    return msg[:4096]


def build_telegram_trends_message(trends: list[dict]) -> str:
    lines = ["🔥 *Tendencias del día — Conexión Summit*\n"]
    for t in trends[:3]:
        keyword = t.get("keyword", "")
        platform = t.get("platform", "")
        how_to_apply = t.get("how_to_apply", "")
        post_idea = t.get("post_idea", "")
        lines.append(f"• *{keyword}* ({platform})\n  {how_to_apply}")
        if post_idea:
            lines.append(f"  💡 {post_idea}")
        lines.append("")
    msg = "\n".join(lines)
    return msg[:4096]


def build_telegram_weekly_summary(metrics_summary: list[dict]) -> str:
    lines = ["📊 *Resumen semanal — Conexión Summit*\n"]
    for m in metrics_summary:
        lines.append(
            f"• {m.get('platform','')}: {m.get('followers',0):,} seguidores · {m.get('engagement_rate',0)}% engagement"
        )
    lines.append("\nVer informe completo en tu email.")
    return "\n".join(lines)[:4096]


# ── Send functions ─────────────────────────────────────────────────────────────

def send_email(config: dict, subject: str, html_body: str) -> bool:
    smtp_host = config.get("smtp_host", "smtp.gmail.com")
    smtp_port = int(config.get("smtp_port", 587))
    sender = config.get("email_sender", "")
    password = config.get("email_password", "")
    recipient = config.get("email_recipient", "")

    if not sender or not password or not recipient:
        logger.warning("Email not configured — skipping send")
        return False

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = recipient
        msg.attach(MIMEText(html_body, "html", "utf-8"))

        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(sender, password)
            server.sendmail(sender, recipient, msg.as_string())
        return True
    except Exception as e:
        logger.warning(f"Email send failed: {e}")
        return False


def send_telegram(config: dict, message: str) -> bool:
    token = config.get("telegram_bot_token", "")
    chat_id = config.get("telegram_chat_id", "")

    if not token or not chat_id:
        logger.warning("Telegram not configured — skipping send")
        return False

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = httpx.post(
            url,
            json={"chat_id": chat_id, "text": message, "parse_mode": "Markdown"},
            timeout=15,
        )
        return resp.status_code == 200
    except Exception as e:
        logger.warning(f"Telegram send failed: {e}")
        return False


# ── Scheduled job implementations ──────────────────────────────────────────────

def run_daily_intelligence_telegram(conn: sqlite3.Connection, config: dict) -> None:
    from backend.services.intelligence import get_articles
    articles = get_articles(conn, limit=5)
    msg = build_telegram_intelligence_message(articles)
    ok = send_telegram(config, msg)
    log_report(conn, "daily_telegram_intelligence", "telegram", "sent" if ok else "skipped")


def run_daily_trends_telegram(conn: sqlite3.Connection, config: dict) -> None:
    from backend.services.trends import get_trends
    trends = get_trends(conn, limit=3)
    msg = build_telegram_trends_message(trends)
    ok = send_telegram(config, msg)
    log_report(conn, "daily_telegram_trends", "telegram", "sent" if ok else "skipped")


def run_daily_email_job(conn: sqlite3.Connection, config: dict, openai_client: Any = None) -> None:
    from backend.services.intelligence import get_articles
    from backend.services.trends import get_trends
    from backend.services.analytics import get_weekly_summary, detect_anomaly

    articles = get_articles(conn, limit=5)
    trends = get_trends(conn, limit=1)
    summary = get_weekly_summary(conn)
    threshold = config.get("alert_threshold_pct", 20)
    anomalies = []
    for row in summary:
        result = detect_anomaly(conn, row["platform"], threshold)
        if result["has_anomaly"]:
            anomalies.append({"platform": row["platform"], **result})

    tip = ""
    if openai_client:
        try:
            resp = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": "Dame un tip de contenido de 1 línea para un CM de emprendimiento LATAM. Solo el tip, sin preámbulo."}],
                max_tokens=80,
                temperature=0.7,
            )
            tip = resp.choices[0].message.content or ""
        except Exception:
            pass

    html = build_daily_email(articles, trends, anomalies, tip)
    ok = send_email(config, "📊 Resumen diario — CM Pro", html)
    log_report(conn, "daily_email", "email", "sent" if ok else "skipped")


def run_weekly_email_job(conn: sqlite3.Connection, config: dict, openai_client: Any = None) -> None:
    from backend.services.intelligence import get_articles
    from backend.services.trends import get_trends
    from backend.services.analytics import get_weekly_summary, detect_anomaly
    from backend.services.planner import get_proposals

    articles = get_articles(conn, limit=7)
    trends = get_trends(conn, limit=5)
    metrics_summary = get_weekly_summary(conn)
    proposals = get_proposals(conn, status="approved", limit=7)
    threshold = config.get("alert_threshold_pct", 20)
    anomalies = []
    for row in metrics_summary:
        result = detect_anomaly(conn, row["platform"], threshold)
        if result["has_anomaly"]:
            anomalies.append({"platform": row["platform"], **result})

    recommendations = ""
    if openai_client and metrics_summary:
        try:
            metrics_text = "; ".join(f"{m['platform']}: {m.get('engagement_rate',0)}% engagement" for m in metrics_summary)
            resp = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": f"Basado en estas métricas ({metrics_text}), da 2-3 recomendaciones estratégicas de contenido para Conexión Summit. Breve y accionable."}],
                max_tokens=200,
                temperature=0.5,
            )
            recommendations = resp.choices[0].message.content or ""
        except Exception:
            pass

    html = build_weekly_email(articles, trends, metrics_summary, proposals, anomalies, recommendations)
    ok = send_email(config, "📈 Informe semanal — CM Pro", html)
    log_report(conn, "weekly_email", "email", "sent" if ok else "skipped")


def run_weekly_telegram_job(conn: sqlite3.Connection, config: dict) -> None:
    from backend.services.analytics import get_weekly_summary
    metrics_summary = get_weekly_summary(conn)
    msg = build_telegram_weekly_summary(metrics_summary)
    ok = send_telegram(config, msg)
    log_report(conn, "weekly_telegram", "telegram", "sent" if ok else "skipped")


# ── Monthly Excel report ───────────────────────────────────────────────────────

def run_monthly_report(conn: sqlite3.Connection, config: dict) -> bytes:
    """Genera Excel del mes pasado con todos los datos relevantes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    today = datetime.now()
    start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    end = today.replace(day=1)
    start_iso = start.isoformat()
    end_iso = end.isoformat()
    month_label = start.strftime("%B %Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")

    def write_header(sheet, row: int, headers: list):
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

    ws["A1"] = f"Reporte mensual — {month_label}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Período"
    ws["B3"] = f"{start.strftime('%Y-%m-%d')} a {end.strftime('%Y-%m-%d')}"

    counts = {}
    counts["articulos_guardados"] = conn.execute(
        "SELECT COUNT(*) FROM saved_items WHERE item_type='article' AND saved_at >= ? AND saved_at < ?",
        (start_iso, end_iso)
    ).fetchone()[0]
    counts["tendencias_guardadas"] = conn.execute(
        "SELECT COUNT(*) FROM saved_items WHERE item_type='trend' AND saved_at >= ? AND saved_at < ?",
        (start_iso, end_iso)
    ).fetchone()[0]
    counts["propuestas_publicadas"] = conn.execute(
        "SELECT COUNT(*) FROM content_proposals WHERE status='published' AND created_at >= ? AND created_at < ?",
        (start_iso, end_iso)
    ).fetchone()[0]
    counts["propuestas_aprobadas"] = conn.execute(
        "SELECT COUNT(*) FROM content_proposals WHERE status='approved' AND created_at >= ? AND created_at < ?",
        (start_iso, end_iso)
    ).fetchone()[0]
    counts["eventos"] = conn.execute(
        "SELECT COUNT(*) FROM events WHERE date >= ? AND date < ?",
        (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))
    ).fetchone()[0]
    ai_cost = conn.execute(
        "SELECT SUM(cost_usd) FROM ai_usage_log WHERE created_at >= ? AND created_at < ?",
        (start_iso, end_iso)
    ).fetchone()[0] or 0

    write_header(ws, 5, ["Métrica", "Valor"])
    ws["A6"] = "Artículos guardados"; ws["B6"] = counts["articulos_guardados"]
    ws["A7"] = "Tendencias guardadas"; ws["B7"] = counts["tendencias_guardadas"]
    ws["A8"] = "Propuestas publicadas"; ws["B8"] = counts["propuestas_publicadas"]
    ws["A9"] = "Propuestas aprobadas"; ws["B9"] = counts["propuestas_aprobadas"]
    ws["A10"] = "Eventos del mes"; ws["B10"] = counts["eventos"]
    ws["A11"] = "Costo IA (USD)"; ws["B11"] = round(ai_cost, 2)

    sh = wb.create_sheet("Artículos guardados")
    write_header(sh, 1, ["Fecha guardado", "Título", "Fuente", "Categoría", "URL", "Resumen"])
    for i, r in enumerate(conn.execute(
        """SELECT saved_at, title, source, category, url, summary FROM saved_items
           WHERE item_type='article' AND saved_at >= ? AND saved_at < ? ORDER BY saved_at DESC""",
        (start_iso, end_iso),
    ).fetchall(), start=2):
        for c, v in enumerate(r, 1):
            sh.cell(row=i, column=c, value=v)

    sh = wb.create_sheet("Tendencias guardadas")
    write_header(sh, 1, ["Fecha guardado", "Tendencia", "Plataforma", "URL fuente", "Descripción"])
    for i, r in enumerate(conn.execute(
        """SELECT saved_at, title, platform, url, summary FROM saved_items
           WHERE item_type='trend' AND saved_at >= ? AND saved_at < ? ORDER BY saved_at DESC""",
        (start_iso, end_iso),
    ).fetchall(), start=2):
        for c, v in enumerate(r, 1):
            sh.cell(row=i, column=c, value=v)

    sh = wb.create_sheet("Propuestas")
    write_header(sh, 1, ["Fecha", "Topic", "Plataforma", "Formato", "Status", "Caption"])
    for i, r in enumerate(conn.execute(
        """SELECT suggested_date, topic, platform, format, status, caption_draft FROM content_proposals
           WHERE created_at >= ? AND created_at < ? ORDER BY suggested_date""",
        (start_iso, end_iso),
    ).fetchall(), start=2):
        for c, v in enumerate(r, 1):
            sh.cell(row=i, column=c, value=v)

    sh = wb.create_sheet("Eventos")
    write_header(sh, 1, ["Fecha", "Título", "Tipo", "Descripción"])
    for i, r in enumerate(conn.execute(
        """SELECT date, title, event_type, description FROM events WHERE date >= ? AND date < ? ORDER BY date""",
        (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')),
    ).fetchall(), start=2):
        for c, v in enumerate(r, 1):
            sh.cell(row=i, column=c, value=v)

    sh = wb.create_sheet("Métricas")
    write_header(sh, 1, ["Plataforma", "Semana", "Followers", "Reach", "Engagement %"])
    for i, r in enumerate(conn.execute(
        """SELECT platform, week_label, followers, reach, engagement_rate FROM metrics
           WHERE recorded_at >= ? AND recorded_at < ? ORDER BY platform, recorded_at""",
        (start_iso, end_iso),
    ).fetchall(), start=2):
        for c, v in enumerate(r, 1):
            sh.cell(row=i, column=c, value=v)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = max_len + 2

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def send_monthly_report_email(conn: sqlite3.Connection, config: dict) -> dict:
    """Genera y envía el reporte mensual por email con Excel adjunto."""
    if not (config.get("email_sender") and config.get("email_recipient")):
        return {"status": "skipped", "reason": "email no configurado"}

    excel_bytes = run_monthly_report(conn, config)
    today = datetime.now()
    last_month = (today.replace(day=1) - timedelta(days=1))
    filename = f"reporte-cm-pro-{last_month.strftime('%Y-%m')}.xlsx"

    try:
        smtp_host = config.get("smtp_host", "smtp.gmail.com")
        smtp_port = int(config.get("smtp_port", 587))
        sender = config["email_sender"]
        password = config.get("email_password", config.get("email_sender_password", ""))
        recipient = config["email_recipient"]

        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = recipient
        msg["Subject"] = f"Reporte mensual CM Pro — {last_month.strftime('%B %Y')}"
        body = (
            f"Hola,\n\n"
            f"Aquí está el reporte mensual de CM Pro para {last_month.strftime('%B %Y')}.\n\n"
            f"Incluye:\n"
            f"- Artículos guardados de Inteligencia\n"
            f"- Tendencias guardadas\n"
            f"- Propuestas publicadas y aprobadas\n"
            f"- Eventos del mes\n"
            f"- Métricas de redes sociales\n"
            f"- Costo de IA del mes\n\n"
            f"Adjunto encontrarás el Excel completo.\n\n"
            f"— CM Pro\n"
        )
        msg.attach(MIMEText(body, "plain"))
        attach = MIMEApplication(excel_bytes, _subtype="xlsx")
        attach.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(attach)

        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as smtp:
            smtp.starttls()
            smtp.login(sender, password)
            smtp.send_message(msg)

        log_report(conn, "monthly", "email", "ok")
        return {"status": "ok", "filename": filename}
    except Exception as e:
        logger.warning(f"send_monthly_report_email failed: {e}")
        log_report(conn, "monthly", "email", "error", str(e))
        return {"status": "error", "error": str(e)}


# ── Weekly intelligence digest ─────────────────────────────────────────────────

def run_weekly_intelligence_email(conn: sqlite3.Connection, config: dict) -> dict:
    """Email con top N artículos (score >= 7) de la última semana."""
    if not (config.get("email_sender") and config.get("email_recipient")):
        return {"status": "skipped", "reason": "email no configurado"}

    n = config.get("count_weekly_top_articles", 5)
    rows = conn.execute(
        """SELECT title_es, title, source, summary, relevance, relevance_score, url, fetched_at
           FROM articles
           WHERE fetched_at >= datetime('now', '-7 days') AND relevance_score >= 7
           ORDER BY relevance_score DESC, fetched_at DESC LIMIT ?""",
        (n,),
    ).fetchall()

    if not rows:
        return {"status": "skipped", "reason": "no hay artículos relevantes esta semana"}

    body = "<h2>Top artículos de la semana — CM Pro</h2><p>Estos son los más relevantes (score &ge; 7):</p>"
    for r in rows:
        title = r[0] or r[1]
        body += (
            f"<div style='border-left:3px solid #6366f1;padding:8px 12px;margin:10px 0;background:#f9fafb;'>"
            f"<h3 style='margin:0 0 4px 0;'>{title}</h3>"
            f"<p style='font-size:12px;color:#6b7280;margin:0 0 6px 0;'>"
            f"{r[2]} &middot; Score {r[5]}/10 &middot; {r[7][:10] if r[7] else ''}</p>"
            f"<p style='font-size:13px;margin:0 0 4px 0;'>{r[3] or ''}</p>"
            f"<p style='font-size:13px;color:#16a34a;margin:0 0 6px 0;'><strong>Relevancia:</strong> {r[4] or ''}</p>"
            f"<a href='{r[6]}' style='font-size:12px;color:#6366f1;'>Leer artículo &rarr;</a>"
            f"</div>"
        )

    try:
        smtp_host = config.get("smtp_host", "smtp.gmail.com")
        smtp_port = int(config.get("smtp_port", 587))
        sender = config["email_sender"]
        password = config.get("email_password", config.get("email_sender_password", ""))
        recipient = config["email_recipient"]

        msg = MIMEMultipart("alternative")
        msg["From"] = sender
        msg["To"] = recipient
        msg["Subject"] = "Top artículos de la semana — CM Pro"
        msg.attach(MIMEText(body, "html"))

        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as smtp:
            smtp.starttls()
            smtp.login(sender, password)
            smtp.send_message(msg)

        log_report(conn, "weekly_intelligence", "email", "ok")
        return {"status": "ok", "articles": len(rows)}
    except Exception as e:
        logger.warning(f"weekly intelligence email failed: {e}")
        log_report(conn, "weekly_intelligence", "email", "error", str(e))
        return {"status": "error", "error": str(e)}
